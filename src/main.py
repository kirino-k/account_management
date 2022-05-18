#########################
##### Version 0.3.0 #####
#########################

import os
import re
import datetime
import openpyxl
import numpy as np
import random
import string
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from openpyxl.styles import PatternFill


dir_name = os.path.abspath(os.path.dirname(__file__))
ledger_dir_name = os.path.join(dir_name, "..")

def ledger_book_button_clicked():
    file_path = filedialog.askopenfilename(
        initialdir=ledger_dir_name, filetypes=[("Excel Files", ".xlsx")]
    )
    if file_path:
        ledger_book_name.set(file_path)


def input_book_buttun_clicked():
    file_path = filedialog.askopenfilename(
        initialdir=dir_name, filetypes=[("Excel Files", ".xlsx")]
    )
    if file_path:
        input_book_name.set(file_path)


def run_button_clicked():
    try:
        # 台帳ファイルと申請フォームの読み込み
        ledger_book = openpyxl.load_workbook(ledger_book_name.get())
        input_book = openpyxl.load_workbook(input_book_name.get())
        if (
            np.count_nonzero(
                list(
                    map(
                        lambda x: x == "プロジェクト" or x == "ユーザーアカウント",
                        ledger_book.sheetnames,
                    )
                )
            )
            != 2
        ):
            messagebox.showerror("エラー", "台帳ファイルのフォーマットが規格と異なります")
        elif (
            np.count_nonzero(
                list(
                    map(
                        lambda x: x == "お読みください" or x == "基本事項の入力" or x == "ユーザーアカウントの入力",
                        input_book.sheetnames,
                    )
                )
            )
            != 3
        ):
            messagebox.showerror("エラー", "申請フォームのフォーマットが規格と異なります")
        else:
            # 台帳ファイルからシートを読み込む
            project_ledger = ledger_book["プロジェクト"]
            account_ledger = ledger_book["ユーザーアカウント"]
            # ユーザーアカウント台帳の内容を取得
            ledger_values = []
            for row in account_ledger.iter_rows(min_row=2):
                ledger_value = [col.value for col in row]
                if ledger_value[0] is None:
                    break
                ledger_values.append(ledger_value)
            # 入力ファイルからシートを読み込む
            project_info = input_book["基本事項の入力"]
            account_info = input_book["ユーザーアカウントの入力"]
            # init
            fill_error = 0
            value_error = 0
            message = ""
            # 基本情報の入力情報を取得
            project_id = project_info.cell(row=3, column=1).value
            project_name = project_info.cell(row=3, column=2).value
            # REDCap研究番号が未記載
            if project_id is None:
                fill_error += 1
                project_info.cell(row=3, column=1).fill = PatternFill(
                    patternType="solid", fgColor="FAD000"
                )
                message += "REDCap研究番号が未記載です。\n"
            elif (
                project_id is not None
                and re.match(r"^(RED)_?[0-9]{4}", project_id.strip()) is None
            ):
                value_error += 1
                project_info.cell(row=3, column=1).fill = PatternFill(
                    patternType="solid", fgColor="FAD000"
                )
                message += "REDCap研究番号を正しく入力してください。\n"
            # 研究課題名名が未記載
            if project_name is None:
                fill_error += 1
                project_info.cell(row=3, column=2).fill = PatternFill(
                    patternType="solid", fgColor="FAD000"
                )
                message += "研究課題名が未記載です。\n"
            # ユーザーアカウントの入力情報を取得
            input_values = []
            last_row = 2
            for row in account_info.iter_rows(min_row=3):
                value = [col.value for col in row]
                # ２セル以上の記入がある列は情報読み込み, 2セル未満の書き込みの場合はループを出る
                if np.count_nonzero(list(map(lambda x: x is not None, value))) < 2:
                    break
                else:
                    last_row += 1
                    # 申請区分が空欄
                    if value[0] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=1).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"  # yellow
                        )
                        message += "行番号 " + str(row[0].row) + ": 「申請区分」が未記載です。\n"
                    # 申請区分が「新規」または「継続」でない
                    elif (
                        value[0] is not None
                        and value[0].strip() != "新規"
                        and value[0].strip() != "継続"
                    ):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=1).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 " + str(row[0].row) + ": 「申請区分」は「新規」または「継続」をご記載ください。\n"
                        )
                    # 希望ユーザー名が空欄
                    if value[1] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=2).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「希望ユーザー名」が未記載です。\n"
                    # 希望ユーザー名に半角英数字, - (ハイフン), _ (アンダースコア) 以外の文字が含まれる
                    elif (
                        value[1] is not None
                        and re.match(r"^[a-zA-Z0-9_-]+$", value[1].strip()) is None
                    ):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=2).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 " + str(row[0].row) + ": 「希望ユーザー名」に使用できない文字が含まれています。\n"
                        )
                    # 希望ユーザー名が10文字以内でない
                    elif value[1] is not None and len(value[1].strip()) > 10:
                        value_error += 1
                        account_info.cell(row=row[0].row, column=2).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「希望ユーザー名」が10文字を超えています。\n"
                    # 氏 (日本語) が空欄
                    if value[2] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=3).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「氏 (日本語)」が未記載です。\n"
                    # 名 (日本語) が空欄
                    if value[3] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=4).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「名 (日本語)」が未記載です。\n"
                    # First_Name (半角英字) が空欄
                    if value[4] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=5).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 " + str(row[0].row) + ": 「First_Name (半角英字)」が未記載です。\n"
                        )
                    # First_Name (半角英字) が半角英字かつ1文字目大文字でない
                    elif (
                        value[4] is not None
                        and re.match(r"^[A-Z]{1}[a-z]+$", value[4].strip()) is None
                    ):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=5).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 "
                            + str(row[0].row)
                            + ": 「First_Name (半角英字)」は半角英字で、1文字目のみ大文字でご記載ください。\n"
                        )
                    # Last_Name (半角英字) が空欄
                    if value[5] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=6).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 " + str(row[0].row) + ": 「Last_Name (半角英字)」が未記載です。\n"
                        )
                    # Last_Name (半角英字) が半角英字かつ1文字目大文字でない
                    elif (
                        value[5] is not None
                        and re.match(r"^[A-Z]{1}[a-z]+$", value[5].strip()) is None
                    ):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=6).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 "
                            + str(row[0].row)
                            + ": 「Last_Name (半角英字)」は半角英字で、1文字目のみ大文字でご記載ください。\n"
                        )
                    # メールアドレスが空欄
                    if value[6] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=7).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「メールアドレス」が未記載です。\n"
                    # メールアドレスがメールアドレスの正規表現に合致しない
                    elif (
                        value[6] is not None
                        and re.match(
                            r"^[A-Za-z0-9]{1}[A-Za-z0-9_.+-]*@{1}[A-Za-z0-9_.+-]+.[A-Za-z0-9]+$",
                            value[6].strip(),
                        )
                        is None
                    ):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=7).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「メールアドレス」を正しくご記入ください。\n"
                    # 所属施設①が空欄
                    if value[7] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=8).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「所属施設①」が未記載です。\n"
                    # その他の施設を選択かつ所属施設②が空欄
                    if value[7] == "その他の施設" and value[8] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=9).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 "
                            + str(row[0].row)
                            + ": 「所属施設②」が未記載です (国立成育医療研究センター以外のにご所属の場合は必ずご記載ください)。\n"
                        )
                    # 所属部署が空欄
                    if value[9] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=10).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「所属部署」が未記載です。\n"
                    # トレーニング受講日が空欄
                    if value[10] is None:
                        fill_error += 1
                        account_info.cell(row=row[0].row, column=11).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += "行番号 " + str(row[0].row) + ": 「トレーニング受講日」が未記載です。\n"
                    # トレーニング受講日が空欄
                    elif value[10] is not None and not isinstance(value[10], datetime.date):
                        value_error += 1
                        account_info.cell(row=row[0].row, column=11).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 "
                            + str(row[0].row)
                            + ": 「トレーニング受講日」の記載が指定した日付フォーマットでの記載になっていません。\n"
                        )
                    # その他の施設を選択かつプロジェクト作成コピー作成の申請権限希望
                    if value[7] == "その他の施設" and value[12] == "〇":
                        value_error += 1
                        account_info.cell(row=row[0].row, column=13).fill = PatternFill(
                            patternType="solid", fgColor="FAD000"
                        )
                        message += (
                            "行番号 "
                            + str(row[0].row)
                            + ": 「プロジェクト作成・コピー作成の申請権限」は国立成育医療研究センター職員以外にはに付与できません。\n"
                        )
                    # 継続申請で、既存アカウントに該当するものがない
                    if (
                        value[0] == "継続"
                        and value[1] is not None
                        and value[4] is not None
                        and value[5] is not None
                        and value[6] is not None
                    ):
                        compatible_account = 0
                        for l in account_ledger.iter_rows():
                            ledger_value = [c.value for c in l]
                            if (
                                ledger_value[2] == value[1].strip()
                                and ledger_value[3] == value[4].strip()
                                and ledger_value[4] == value[5].strip()
                                and ledger_value[5] == value[6].strip()
                            ):
                                compatible_account += 1
                        if compatible_account == 0:
                            value_error += 1
                            message += (
                                "行番号 "
                                + str(row[0].row)
                                + ": 「継続」の申請となっておりますが、記入内容に一致する既存アカウントがみつかりません。\n"
                            )
                            for i in range(1, 8):
                                account_info.cell(
                                    row=row[0].row, column=i
                                ).fill = PatternFill(patternType="solid", fgColor="FAD000")
                input_values.append(value)

            is_to_contact = [int(value[11] == "〇") for value in input_values]
            # 連絡担当者の列の〇が0個
            if np.count_nonzero(is_to_contact) == 0:
                value_error += 1
                for i in range(3, last_row + 1):
                    account_info.cell(row=i, column=12).fill = PatternFill(
                        patternType="solid", fgColor="FAD000"
                    )
                message += "連絡担当者を1名お選びください。\n"
            # 連絡担当者の列の〇が2個以上
            elif np.count_nonzero(is_to_contact) > 1:
                value_error += 1
                for i in range(3, last_row + 1):
                    account_info.cell(row=i, column=12).fill = PatternFill(
                        patternType="solid", fgColor="FAD000"
                    )
                message += "連絡担当者が複数選択されています。1名にしてください。\n"
            apply_for_create = [int(value[12] == "〇") for value in input_values]
            # 申請権限の付与希望の列の〇が0個
            if np.count_nonzero(apply_for_create) == 0:
                value_error += 1
                for i in range(3, last_row + 1):
                    account_info.cell(row=i, column=13).fill = PatternFill(
                        patternType="solid", fgColor="FAD000"
                    )
                message += "プロジェクト作成・コピー作成の申請権限を1名に付与します。希望者を1名お選びください。\n"
            # 申請権限の付与希望の列の〇が2個以上
            if np.count_nonzero(apply_for_create) > 1:
                value_error += 1
                for i in range(3, last_row + 1):
                    account_info.cell(row=i, column=13).fill = PatternFill(
                        patternType="solid", fgColor="FAD000"
                    )
                message += "プロジェクト作成・コピー作成の申請権限の希望者が複数選択されています。1名にしてください。\n"
            # 記入に不備または問い合わせ項目がある場合
            if fill_error or value_error:
                reject_book_path = (
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_再提出用申請フォーム.xlsx"
                )
                input_book.save(reject_book_path)
                message = (
                    "記入に不備があるため、差し戻しとさせていただきます。\n"
                    + "再提出用申請フォームの黄色ハイライトで示した該当箇所をご修正のうえ、改めてご提出願います。\n\n"
                    + "なお、下記に各該当箇所の指摘事項を記載いたします。ご参照ください。\n\n"
                    + message
                )
                with open(
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_差し戻しメッセージ.txt",
                    "w",
                ) as f:
                    f.write(message)
                    f.close()
                messagebox.showinfo("お知らせ", "記入に不備あり差し戻しです。\n")
                root.quit()
            # 差し戻しなし
            else:
                # 台帳上書き前にアーカイブ用に保存
                archive_ledger_path = (
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_アーカイブ台帳.xlsx"
                )
                ledger_book.save(archive_ledger_path)
                # 台帳ファイルに各種の重複がある場合の事務局修正
                account_names_in_ledger = [value[2] for value in ledger_values]
                csv = "Username,First name,Last name,Email address,Institution ID,Sponsor username,Expiration,Comments\n"
                account_exist = 0
                for index, input_value in enumerate(input_values):
                    account_name = input_value[1].strip()
                    user_name_ja = input_value[2].strip() + " " + input_value[3].strip()
                    if input_value[7] == "その他の施設":
                        institute = input_value[8].strip()
                    else:
                        institute = input_value[7]
                    account_exist_in_row = 0
                    for ledger_value in ledger_values:
                        # ユーザ名、ローマ字氏名、メールアドレスが既存アカウントと同一だが「新規」として申請
                        # ->「継続」に事務局修正
                        if (
                            input_value[0] == "新規"
                            and ledger_value[2] == account_name
                            and ledger_value[3] == input_value[4].strip()
                            and ledger_value[4] == input_value[5].strip()
                            and ledger_value[5] == input_value[6].strip()
                        ):
                            account_exist_in_row += 1
                            account_info.cell(row=index + 3, column=1).value = "継続"
                            account_info.cell(row=index + 3, column=1).fill = PatternFill(
                                patternType="solid", fgColor="FAD000"
                            )
                            message += (
                                "行番号 "
                                + str(index + 3)
                                + ": "
                                + user_name_ja
                                + " (ユーザー名 "
                                + ledger_value[2]
                                + ") さんは既にREDCapに登録されています。「継続」として登録しました。\n"
                            )
                            break
                        # ユーザー名以外が既存アカウントと同一の情報で、「新規」として申請
                        # ->ユーザー名を既存のものにして「継続」に事務局修正
                        elif (
                            input_value[0] == "新規"
                            and ledger_value[2] != account_name
                            and ledger_value[3] == input_value[4].strip()
                            and ledger_value[4] == input_value[5].strip()
                            and ledger_value[5] == input_value[6].strip()
                        ):
                            account_exist_in_row += 1
                            account_info.cell(row=index + 3, column=1).value = "継続"
                            account_info.cell(row=index + 3, column=1).fill = PatternFill(
                                patternType="solid", fgColor="FAD000"
                            )
                            account_info.cell(row=index + 3, column=2).value = ledger_value[
                                2
                            ]
                            account_info.cell(row=index + 3, column=2).fill = PatternFill(
                                patternType="solid", fgColor="FAD000"
                            )
                            account_name = ledger_value[2]
                            message += (
                                "行番号 "
                                + str(index + 3)
                                + ": "
                                + user_name_ja
                                + " さんは既にREDCapに登録されています。今回入力されたユーザー名 ("
                                + input_value[1].strip()
                                + ") ではなく、既存のユーザー名 ("
                                + account_name
                                + ") で「継続」として登録しました。\n"
                            )
                            break
                    # アカウント名のかぶりを解消
                    if (
                        account_exist_in_row == 0
                        and input_value[0] == "新規"
                        and len(account_names_in_ledger) > 0
                    ):
                        while (
                            np.count_nonzero(
                                list(
                                    map(
                                        lambda x: x == account_name,
                                        account_names_in_ledger,
                                    )
                                )
                            )
                            > 0
                        ):
                            account_name += "".join(
                                [
                                    random.choice(string.ascii_letters + string.digits)
                                    for i in range(2)
                                ]
                            )
                        account_names_in_ledger.append(account_name)
                        if account_name != input_value[1].strip():
                            account_exist_in_row += 1
                            account_info.cell(row=index + 3, column=2).value = account_name
                            account_info.cell(row=index + 3, column=2).fill = PatternFill(
                                patternType="solid", fgColor="FAD000"
                            )
                            message += (
                                "行番号 "
                                + str(index + 3)
                                + ": "
                                + user_name_ja
                                + " さんはの希望されたユーザー名は既に使用されています。ユーザー名 ("
                                + account_name
                                + ") として登録しした。\n"
                            )
                    # アカウント台帳に追記
                    account_ledger.append(
                        [
                            datetime.datetime.now(),
                            project_id,
                            account_name,
                            input_value[4].strip(),
                            input_value[5].strip(),
                            input_value[6].strip(),
                            user_name_ja,
                            institute,
                            input_value[9].strip(),
                            input_value[10],
                            int(input_value[12] == "〇"),
                        ]
                    )
                    # REDCapインポート用CSVファイルの内容を追加
                    csv += (
                        account_name
                        + ","
                        + input_value[4].strip()
                        + ","
                        + input_value[5].strip()
                        + ","
                        + input_value[6].strip()
                        + ",,,,\n"
                    )
                    account_exist += account_exist_in_row
                # 連絡担当者情報抽出
                contact_person = input_values[is_to_contact.index(1)]
                # プロジェクト台帳に追記
                project_ledger.append(
                    [
                        datetime.datetime.now(),
                        project_id,
                        project_name,
                        contact_person[2].strip() + " " + contact_person[3].strip(),
                        contact_person[6].strip(),
                    ]
                )
                # 台帳上書き保存
                ledger_book.save(ledger_book_name.get())
                # 登録済み申請フォームを作成
                account_info.cell(1, 2).value = "確定ユーザー名"
                processed_book_path = (
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_登録済み申請フォーム.xlsx"
                )
                input_book.save(processed_book_path)
                with open(
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_import.csv",
                    "w",
                ) as f:
                    f.write(csv)
                    f.close()
                if account_exist:
                    message = (
                        "ユーザー登録が終了しました。\n"
                        + "登録内容を記載したエクセルファイルをご確認ください。\n\n"
                        + "なお、事務局修正を行った箇所があります。\n"
                        + "エクセルファイルの黄色ハイライトで示した該当箇所および下記のお知らせをご参照ください。\n\n"
                        + message
                    )
                else:
                    message = "ユーザー登録が終了しました。\n登録内容を記載したエクセルファイルをご確認ください。\n\n" + message
                with open(
                    project_id
                    + "_"
                    + str(datetime.datetime.now().strftime("%Y%m%d"))
                    + "_登録終了メッセージ.txt",
                    "w",
                ) as f:
                    f.write(message)
                    f.close()
                messagebox.showinfo("お知らせ", "処理が終わりました。\n")
                root.quit()

    except:
        messagebox.showerror("エラー", "うまく処理が終了しませんでした。\n")


if __name__ == "__main__":
    root = tk.Tk()
    root.title("REDCapアカウント管理お助けツール")
    root.geometry("600x150")

    ledger_book_frame = ttk.Frame(root, padding=10)
    ledger_book_frame.grid()

    ledger_book_set = tk.StringVar()
    ledger_book_set.set("台帳ファイル: ")
    ledger_book_label = ttk.Label(ledger_book_frame, textvariable=ledger_book_set)
    ledger_book_label.grid(row=0, column=0)

    ledger_book_name = tk.StringVar()
    ledger_book_entry = ttk.Entry(
        ledger_book_frame, textvariable=ledger_book_name, width=50
    )
    ledger_book_entry.grid(row=0, column=1)

    ledger_book_button = ttk.Button(
        ledger_book_frame, text="参照", command=ledger_book_button_clicked
    )
    ledger_book_button.grid(row=0, column=2)

    input_book_frame = ttk.Frame(root, padding=10)
    input_book_frame.grid()

    input_book_set = tk.StringVar()
    input_book_set.set("申請フォーム: ")
    input_book_label = ttk.Label(input_book_frame, textvariable=input_book_set)
    input_book_label.grid(row=0, column=0)

    input_book_name = tk.StringVar()
    input_book_entry = ttk.Entry(
        input_book_frame, textvariable=input_book_name, width=50
    )
    input_book_entry.grid(row=0, column=1)

    input_book_button = ttk.Button(
        input_book_frame, text="参照", command=input_book_buttun_clicked
    )
    input_book_button.grid(row=0, column=2)

    run_frame = ttk.Frame(root, padding=10)
    run_frame.grid()

    run_button = ttk.Button(run_frame, text="実行", command=run_button_clicked)
    run_button.grid(row=0, column=0)

    root.mainloop()
